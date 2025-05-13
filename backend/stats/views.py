from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import FilterRequest
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from collections import defaultdict
import statistics
from datetime import datetime


def export_filters_to_excel(request):
    # Получаем параметры дат из запроса
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Преобразуем строки в даты с учетом временной зоны
    try:
        if start_date:
            naive_start = datetime.strptime(start_date, '%Y-%m-%d')
            start_date = timezone.make_aware(naive_start)
        if end_date:
            naive_end = datetime.strptime(end_date, '%Y-%m-%d')
            end_date = timezone.make_aware(naive_end)
    except ValueError:
        start_date = end_date = None

    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filter_stats_{}.xlsx"'.format(
        timezone.now().strftime('%Y-%m-%d_%H-%M-%S')
    )

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Фильтруем запросы по дате, если указаны даты
    filter_requests = FilterRequest.objects.all()
    if start_date:
        filter_requests = filter_requests.filter(created_at__gte=start_date)
    if end_date:
        filter_requests = filter_requests.filter(created_at__lte=end_date)
    
    if not filter_requests:
        return HttpResponse("Нет данных для экспорта", status=400)

    # --- Лист 1: Данные ---
    ws_data = wb.create_sheet(title="Данные")
    headers = ['Дата запроса']

    districts = ['Кировский', 'Ленинский', 'Советский', 'Трусовский']
    rooms = ['1', '2', '3', '4', '5+']

    for district in districts:
        headers.append(f"Район: {district}")
    for room in rooms:
        headers.append(f"Комнат: {room}")

    headers.extend([
        'Цена от', 'Цена до',
        'Цена за м² от', 'Цена за м² до',
        'Общая площадь от', 'Общая площадь до',
        'Этаж от', 'Этаж до',
        'Не первый этаж', 'Не последний этаж',
        'Санузел: Разделенный', 'Санузел: Совмещенный',
        'Ремонт: Евро', 'Ремонт: Косметический', 'Ремонт: Без ремонта',
        'Площадь кухни от', 'Площадь кухни до',
        'Высота потолков от', 'Высота потолков до',
        'Окна: На улицу', 'Окна: Во двор',
        'Балкон', 'Лоджия',
        'Год постройки от', 'Год постройки до',
        'Этажей в доме от', 'Этажей в доме до',
        'Тип дома: Панельный', 'Тип дома: Кирпичный', 'Тип дома: Монолитный',
        'Лифт: Пассажирский', 'Лифт: Грузовой',
        'Парковка: Подземная', 'Парковка: Наземная',
        'Близость остановок',
        'Близость магазина продуктов',
        'Близость ресторанов',
        'Близость школы',
        'Близость детского сада',
        'Близость пункта выдачи',
        'Близость поликлиники',
        'Близость аптеки',
        'Близость банка',
        'Близость центра',
        'Близость тренажерного зала',
        'Близость ТЦ',
        'Близость ВУЗа',
        'Близость салона красоты',
        'Близость религиозных объектов',
    ])

    ws_data.append(headers)

    # --- Статистика ---
    stats = {
        'total': 0,
        'districts': defaultdict(int),
        'rooms': defaultdict(int),
        'price_min': [],
        'price_max': [],
        'area_min': [],
        'area_max': [],
        'repair_types': defaultdict(int),
    }

    def translate_proximity(value):
        if value == 'any':
            return 'Неважно'
        elif value == 'close':
            return 'Рядом'
        elif value == 'far':
            return 'Подальше'
        return value or ''

    for req in filter_requests:
        filters = req.filters or {}
        stats['total'] += 1

        row = [req.created_at.strftime('%Y-%m-%d %H:%M:%S')]

        # Районы
        for district in districts:
            val = 'Да' if district in filters.get('districts', []) else 'Нет'
            if val == 'Да':
                stats['districts'][district] += 1
            row.append(val)

        # Комнаты
        for room in rooms:
            room_value = str(room) if room != '5+' else '5+'
            val = 'Да' if room_value in map(str, filters.get('rooms', [])) else 'Нет'
            if val == 'Да':
                stats['rooms'][room_value] += 1
            row.append(val)

        # Основные параметры
        price_min = filters.get('priceMin')
        price_max = filters.get('priceMax')

        if price_min:
            try:
                stats['price_min'].append(float(price_min))
            except ValueError:
                pass
        if price_max:
            try:
                stats['price_max'].append(float(price_max))
            except ValueError:
                pass
            
        area_min = filters.get('totalAreaMin')
        area_max = filters.get('totalAreaMax')

        if area_min:
            try:
                stats['area_min'].append(float(area_min))
            except ValueError:
                pass
        if area_max:
            try:
                stats['area_max'].append(float(area_max))
            except ValueError:
                pass

        row.extend([
            price_min or '',
            price_max or '',
            filters.get('pricePerMeterMin') or '',
            filters.get('pricePerMeterMax') or '',
            filters.get('totalAreaMin') or '',
            filters.get('totalAreaMax') or '',
            filters.get('floorMin') or '',
            filters.get('floorMax') or '',
            'Да' if filters.get('notFirstFloor') else 'Нет',
            'Да' if filters.get('notLastFloor') else 'Нет',
        ])

        # Санузел
        for t in ['Разделенный', 'Совмещенный']:
            val = 'Да' if t in filters.get('bathroomType', []) else 'Нет'
            row.append(val)

        # Ремонт
        for t in ['Евро', 'Косметический', 'Без ремонта']:
            val = 'Да' if t in filters.get('repairType', []) else 'Нет'
            if val == 'Да':
                stats['repair_types'][t] += 1
            row.append(val)
        
        # Площадь кухни
        row.extend([
            filters.get('kitchenAreaMin') or '',
            filters.get('kitchenAreaMax') or '',

            # Высота потолков
            filters.get('ceilingHeightMin') or '',
            filters.get('ceilingHeightMax') or '',

            # Окна
            'Да' if 'На улицу' in filters.get('windows', []) else 'Нет',
            'Да' if 'Во двор' in filters.get('windows', []) else 'Нет',

            # Балкон / Лоджия
            'Да' if 'Балкон' in filters.get('balcony', []) else 'Нет',
            'Да' if 'Лоджия' in filters.get('balcony', []) else 'Нет',

            # Год постройки
            filters.get('yearOfConstructionMin') or '',
            filters.get('yearOfConstructionMax') or '',

            # Этажей в доме
            filters.get('numberOfFloorMin') or '',
            filters.get('numberOfFloorMax') or '',

            # Тип дома
            'Да' if 'Панельный' in filters.get('houseType', []) else 'Нет',
            'Да' if 'Кирпичный' in filters.get('houseType', []) else 'Нет',
            'Да' if 'Монолитный' in filters.get('houseType', []) else 'Нет',

            # Лифт
            'Да' if 'Пассажирский' in filters.get('elevator', []) else 'Нет',
            'Да' if 'Грузовой' in filters.get('elevator', []) else 'Нет',

            # Парковка
            'Да' if 'Подземная' in filters.get('parking', []) else 'Нет',
            'Да' if 'Наземная' in filters.get('parking', []) else 'Нет',
        ])

        # Близость объектов (используем translate_proximity)
        proximity_fields = [
            'stops', 'grocery_store', 'restaurant', 'school', 'kindergarten',
            'pickup_point', 'polyclinic', 'pharmacy', 'bank', 'center',
            'gym', 'mall', 'college_and_university', 'beauty_salon', 'religious'
        ]

        for field in proximity_fields:
            row.append(translate_proximity(filters.get(field)))    

        ws_data.append(row)

    # --- Лист 2: Статистика ---
    ws_stats = wb.create_sheet(title="Статистика")

    # Устанавливаем фиксированные ширины столбцов
    ws_stats.column_dimensions['A'].width = 280 / 7  # 1 пиксель ≈ 1/7 единицы ширины в Excel
    ws_stats.column_dimensions['B'].width = 150 / 7

    # Добавляем заголовок с периодом
    period_title = "Статистика за период: "
    if start_date and end_date:
        period_title += f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
    elif start_date:
        period_title += f"с {start_date.strftime('%d.%m.%Y')}"
    elif end_date:
        period_title += f"по {end_date.strftime('%d.%m.%Y')}"
    else:
        period_title += "за все время"

    # Добавляем заголовок с большим шрифтом
    ws_stats.append([period_title])
    ws_stats.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws_stats.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)

    # Пустая строка для разделения
    ws_stats.append([])
    ws_stats.append(['Метрика', 'Значение'])

    ws_stats.append(['Всего запросов', stats['total']])
    avg_price_min = statistics.mean(stats['price_min']) if stats['price_min'] else 0
    avg_price_max = statistics.mean(stats['price_max']) if stats['price_max'] else 0
    average_price = round((avg_price_min + avg_price_max) / 2) if (avg_price_min or avg_price_max) else 0
    ws_stats.append(['Средняя цена', f"{average_price} ₽"])

    avg_area_min = statistics.mean(stats['area_min']) if stats['area_min'] else 0
    avg_area_max = statistics.mean(stats['area_max']) if stats['area_max'] else 0
    average_area = round((avg_area_min + avg_area_max) / 2) if (avg_area_min or avg_area_max) else 0
    ws_stats.append(['Средняя площадь', f"{average_area} м²"])

    ws_stats.append(['Часто выбираемый район', max(stats['districts'], key=stats['districts'].get)])
    ws_stats.append(['Часто выбираемое количество комнат', max(stats['rooms'], key=stats['rooms'].get)])

    # Форматирование
    for cell in ws_stats[3]:  # Теперь заголовки таблицы на строке 3
        cell.font = Font(bold=True)

    # Автоподгон ширины для остальных листов (если нужно)
    for ws in [ws_data]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


class FilterStatsAPIView(APIView):
    def post(self, request):
        try:
            FilterRequest.objects.create(
                filters=request.data.get('filters', {})
            )
            return Response({'status': 'success'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)